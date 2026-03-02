"""
Excel/CSV Analyzer with Gemini AI
Analyzes tabular data using Gemini 2.5 Flash
"""
from fastapi import APIRouter, HTTPException, UploadFile, File, Depends, Form
from pydantic import BaseModel
from typing import Optional, List
import os
import uuid
import tempfile
import aiofiles
from datetime import datetime, timezone
from dotenv import load_dotenv

load_dotenv()

router = APIRouter(prefix="/analyzer", tags=["analyzer"])

# Store active analysis sessions
analysis_sessions = {}

class AnalyzeRequest(BaseModel):
    session_id: str
    question: str

class AnalysisResponse(BaseModel):
    answer: str
    session_id: str

def setup_analyzer_routes(db, get_current_user):
    """Setup analyzer routes with dependencies"""
    
    EMERGENT_KEY = os.environ.get("EMERGENT_LLM_KEY")
    CLAUDE_KEY = os.environ.get("CLAUDE_API_KEY")
    
    @router.post("/upload")
    async def upload_for_analysis(
        file: UploadFile = File(...),
        current_user: dict = Depends(get_current_user)
    ):
        """Upload Excel/CSV file for analysis"""
        
        # Validate file type
        allowed_types = [
            "text/csv",
            "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            "application/vnd.ms-excel"
        ]
        
        if file.content_type not in allowed_types:
            raise HTTPException(
                status_code=400, 
                detail="Only Excel (.xlsx) and CSV files are supported"
            )
        
        content = await file.read()
        
        # Limit file size (10MB)
        if len(content) > 10 * 1024 * 1024:
            raise HTTPException(status_code=400, detail="File too large (max 10MB)")
        
        # Create session
        session_id = str(uuid.uuid4())
        
        # Save file temporarily
        ext = ".csv" if "csv" in file.content_type else ".xlsx"
        temp_path = f"/tmp/analyzer_{session_id}{ext}"
        
        async with aiofiles.open(temp_path, 'wb') as f:
            await f.write(content)
        
        # Parse file to get preview
        preview_data = []
        total_rows = 0
        columns = []
        
        try:
            if ext == ".csv":
                import csv
                from io import StringIO
                text = content.decode('utf-8', errors='ignore')
                reader = csv.reader(StringIO(text))
                rows = list(reader)
                if rows:
                    columns = rows[0]
                    total_rows = len(rows) - 1
                    preview_data = rows[:11]  # Header + 10 rows
            else:
                from openpyxl import load_workbook
                from io import BytesIO
                wb = load_workbook(BytesIO(content), data_only=True)
                sheet = wb.active
                rows = list(sheet.iter_rows(values_only=True))
                if rows:
                    columns = [str(c) if c else f"Col_{i}" for i, c in enumerate(rows[0])]
                    total_rows = len(rows) - 1
                    for row in rows[:11]:
                        preview_data.append([str(c) if c is not None else "" for c in row])
        except Exception as e:
            raise HTTPException(status_code=400, detail=f"Failed to parse file: {str(e)}")
        
        # Store session info
        mime_type = "text/csv" if ext == ".csv" else "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        analysis_sessions[session_id] = {
            "file_path": temp_path,
            "file_name": file.filename,
            "mime_type": mime_type,
            "user_id": current_user["id"],
            "columns": columns,
            "total_rows": total_rows,
            "created_at": datetime.now(timezone.utc).isoformat(),
            "messages": []
        }
        
        return {
            "session_id": session_id,
            "file_name": file.filename,
            "columns": columns,
            "total_rows": total_rows,
            "preview": preview_data
        }
    
    @router.post("/ask")
    async def ask_about_data(
        request: AnalyzeRequest,
        current_user: dict = Depends(get_current_user)
    ):
        """Ask a question about the uploaded data"""
        
        session = analysis_sessions.get(request.session_id)
        if not session:
            raise HTTPException(status_code=404, detail="Session not found. Please upload a file first.")
        
        if session["user_id"] != current_user["id"]:
            raise HTTPException(status_code=403, detail="Access denied")
        
        # Use Claude API key if available, otherwise fall back to Emergent key
        api_key = CLAUDE_KEY or EMERGENT_KEY
        use_claude = bool(CLAUDE_KEY)
        
        if not api_key:
            raise HTTPException(status_code=500, detail="API key not configured")
        
        try:
            from emergentintegrations.llm.chat import LlmChat, UserMessage
            
            # Read file content as text
            file_text = ""
            max_rows_for_context = 10000  # Increased limit
            
            try:
                if session["mime_type"] == "text/csv":
                    import csv
                    from io import StringIO
                    with open(session["file_path"], 'r', encoding='utf-8', errors='ignore') as f:
                        content = f.read()
                    reader = csv.reader(StringIO(content))
                    rows = list(reader)
                    
                    if rows:
                        headers = rows[0]
                        total_data_rows = len(rows) - 1
                        shown_rows = min(total_data_rows, max_rows_for_context)
                        file_text = f"Cols:{','.join(headers)}\n"
                        
                        # Super compact: R1:v1|v2|v3 (max 150 chars per value)
                        for i, row in enumerate(rows[1:max_rows_for_context+1], 1):
                            vals = [v.strip()[:150] for v in row if v and v.strip()]
                            if vals:
                                file_text += f"R{i}:{'|'.join(vals)}\n"
                else:
                    # Excel file
                    from openpyxl import load_workbook
                    wb = load_workbook(session["file_path"], data_only=True)
                    sheet = wb.active
                    rows = list(sheet.iter_rows(values_only=True))
                    
                    if rows:
                        headers = [str(c) if c else f"C{i}" for i, c in enumerate(rows[0])]
                        total_data_rows = len(rows) - 1
                        shown_rows = min(total_data_rows, max_rows_for_context)
                        file_text = f"Cols:{','.join(headers)}\n"
                        
                        for i, row in enumerate(rows[1:max_rows_for_context+1], 1):
                            vals = [str(c).strip()[:150] for c in row if c is not None and str(c).strip()]
                            if vals:
                                file_text += f"R{i}:{'|'.join(vals)}\n"
            except Exception as read_error:
                raise HTTPException(status_code=500, detail=f"Failed to read file: {str(read_error)}")
            
            # Limit to 8K chars for Claude rate limit (10K tokens/min)
            if len(file_text) > 8000:
                file_text = file_text[:8000] + "\n[TRUNCATED - Claude rate limit]"
            
            # Create chat with Claude for Excel analysis
            chat = LlmChat(
                api_key=CLAUDE_KEY,
                session_id=f"analyzer_{request.session_id}",
                system_message=f"""Data analyst for "{session['file_name']}" ({session['total_rows']} rows).
Cols: {', '.join(session['columns'][:10])}.
Rules: List ALL matches with row numbers. Be concise."""
            ).with_model("anthropic", "claude-sonnet-4-20250514")
            
            # Build message with context from previous messages
            context = ""
            if session["messages"]:
                context = "Previous conversation:\n"
                for msg in session["messages"][-2:]:  # Last 2 messages for context
                    context += f"Q: {msg['question']}\nA: {msg['answer'][:500]}...\n\n"
            
            full_question = f"{context}\nDATA:\n{file_text}\n\nQuestion: {request.question}"
            
            user_message = UserMessage(text=full_question)
            
            # Get response
            response = await chat.send_message(user_message)
            
            # Calculate debug info
            rows_in_context = file_text.count('\nR')
            chars_in_context = len(file_text)
            
            # Add debug info to response
            debug_info = f"\n\n---\n_📊 Debug: {rows_in_context} строк отправлено, {chars_in_context:,} символов_"
            response_with_debug = response + debug_info
            
            # Store in session history (without debug)
            session["messages"].append({
                "question": request.question,
                "answer": response,
                "timestamp": datetime.now(timezone.utc).isoformat()
            })
            
            return {
                "answer": response_with_debug,
                "session_id": request.session_id,
                "debug": {
                    "rows_sent": rows_in_context,
                    "chars_sent": chars_in_context,
                    "total_rows": session["total_rows"],
                    "truncated": chars_in_context >= 300000
                }
            }
            
        except HTTPException:
            raise
        except Exception as e:
            error_msg = str(e)
            if "token" in error_msg.lower() or "context" in error_msg.lower():
                raise HTTPException(status_code=400, detail="Файл слишком большой. Попробуйте с файлом меньшего размера (до 500 строк).")
            raise HTTPException(status_code=500, detail=f"Analysis failed: {error_msg}")
    
    @router.get("/session/{session_id}")
    async def get_session(
        session_id: str,
        current_user: dict = Depends(get_current_user)
    ):
        """Get session info and history"""
        session = analysis_sessions.get(session_id)
        if not session:
            raise HTTPException(status_code=404, detail="Session not found")
        
        if session["user_id"] != current_user["id"]:
            raise HTTPException(status_code=403, detail="Access denied")
        
        return {
            "session_id": session_id,
            "file_name": session["file_name"],
            "columns": session["columns"],
            "total_rows": session["total_rows"],
            "messages": session["messages"],
            "created_at": session["created_at"]
        }
    
    @router.delete("/session/{session_id}")
    async def delete_session(
        session_id: str,
        current_user: dict = Depends(get_current_user)
    ):
        """Delete analysis session"""
        session = analysis_sessions.get(session_id)
        if not session:
            raise HTTPException(status_code=404, detail="Session not found")
        
        if session["user_id"] != current_user["id"]:
            raise HTTPException(status_code=403, detail="Access denied")
        
        # Delete temp file
        try:
            import os
            if os.path.exists(session["file_path"]):
                os.remove(session["file_path"])
        except:
            pass
        
        del analysis_sessions[session_id]
        
        return {"message": "Session deleted"}
    
    @router.get("/session/{session_id}/export/excel")
    async def export_to_excel(
        session_id: str,
        current_user: dict = Depends(get_current_user)
    ):
        """Export analysis session to Excel file"""
        from openpyxl import Workbook
        from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
        from io import BytesIO
        from fastapi.responses import StreamingResponse
        
        session = analysis_sessions.get(session_id)
        if not session:
            raise HTTPException(status_code=404, detail="Session not found")
        
        if session["user_id"] != current_user["id"]:
            raise HTTPException(status_code=403, detail="Access denied")
        
        # Create workbook
        wb = Workbook()
        ws = wb.active
        ws.title = "Analysis Report"
        
        # Styles
        header_font = Font(bold=True, size=14, color="FFFFFF")
        header_fill = PatternFill(start_color="22C55E", end_color="22C55E", fill_type="solid")
        question_fill = PatternFill(start_color="E5E7EB", end_color="E5E7EB", fill_type="solid")
        answer_fill = PatternFill(start_color="F0FDF4", end_color="F0FDF4", fill_type="solid")
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        # Header
        ws.merge_cells('A1:B1')
        ws['A1'] = f"Analysis Report: {session['file_name']}"
        ws['A1'].font = header_font
        ws['A1'].fill = header_fill
        ws['A1'].alignment = Alignment(horizontal='center')
        
        # File info
        ws['A3'] = "File:"
        ws['B3'] = session['file_name']
        ws['A4'] = "Rows:"
        ws['B4'] = session['total_rows']
        ws['A5'] = "Columns:"
        ws['B5'] = ", ".join(session['columns'][:10])
        ws['A6'] = "Created:"
        ws['B6'] = session['created_at'][:19].replace('T', ' ')
        
        for row in range(3, 7):
            ws[f'A{row}'].font = Font(bold=True)
        
        # Q&A Section
        ws['A8'] = "Questions & Answers"
        ws['A8'].font = Font(bold=True, size=12)
        
        current_row = 10
        for i, msg in enumerate(session.get("messages", []), 1):
            # Question
            ws[f'A{current_row}'] = f"Q{i}:"
            ws[f'A{current_row}'].font = Font(bold=True)
            ws[f'A{current_row}'].fill = question_fill
            
            ws.merge_cells(f'B{current_row}:D{current_row}')
            ws[f'B{current_row}'] = msg['question']
            ws[f'B{current_row}'].fill = question_fill
            ws[f'B{current_row}'].alignment = Alignment(wrap_text=True)
            
            current_row += 1
            
            # Answer
            ws[f'A{current_row}'] = f"A{i}:"
            ws[f'A{current_row}'].font = Font(bold=True, color="22C55E")
            ws[f'A{current_row}'].fill = answer_fill
            
            ws.merge_cells(f'B{current_row}:D{current_row}')
            ws[f'B{current_row}'] = msg['answer']
            ws[f'B{current_row}'].fill = answer_fill
            ws[f'B{current_row}'].alignment = Alignment(wrap_text=True)
            
            current_row += 2
        
        # Adjust column widths
        ws.column_dimensions['A'].width = 10
        ws.column_dimensions['B'].width = 60
        ws.column_dimensions['C'].width = 20
        ws.column_dimensions['D'].width = 20
        
        # Save to buffer
        buffer = BytesIO()
        wb.save(buffer)
        buffer.seek(0)
        
        filename = f"analysis_{session['file_name'].rsplit('.', 1)[0]}_{session_id[:8]}.xlsx"
        
        return StreamingResponse(
            buffer,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": f"attachment; filename={filename}"}
        )
    
    @router.get("/session/{session_id}/export/pdf")
    async def export_to_pdf(
        session_id: str,
        current_user: dict = Depends(get_current_user)
    ):
        """Export analysis session to PDF file"""
        from reportlab.lib.pagesizes import A4
        from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
        from reportlab.lib.units import inch
        from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table as PDFTable, TableStyle
        from reportlab.lib import colors
        from reportlab.pdfbase import pdfmetrics
        from reportlab.pdfbase.ttfonts import TTFont
        from io import BytesIO
        from fastapi.responses import StreamingResponse
        
        session = analysis_sessions.get(session_id)
        if not session:
            raise HTTPException(status_code=404, detail="Session not found")
        
        if session["user_id"] != current_user["id"]:
            raise HTTPException(status_code=403, detail="Access denied")
        
        # Create PDF buffer
        buffer = BytesIO()
        doc = SimpleDocTemplate(buffer, pagesize=A4, rightMargin=72, leftMargin=72, topMargin=72, bottomMargin=72)
        
        # Styles
        styles = getSampleStyleSheet()
        title_style = ParagraphStyle(
            'CustomTitle',
            parent=styles['Heading1'],
            fontSize=18,
            spaceAfter=20,
            textColor=colors.HexColor('#22C55E')
        )
        heading_style = ParagraphStyle(
            'CustomHeading',
            parent=styles['Heading2'],
            fontSize=14,
            spaceAfter=10,
            textColor=colors.HexColor('#1F2937')
        )
        question_style = ParagraphStyle(
            'Question',
            parent=styles['Normal'],
            fontSize=11,
            spaceAfter=5,
            textColor=colors.HexColor('#374151'),
            backColor=colors.HexColor('#F3F4F6'),
            borderPadding=5
        )
        answer_style = ParagraphStyle(
            'Answer',
            parent=styles['Normal'],
            fontSize=11,
            spaceAfter=15,
            textColor=colors.HexColor('#1F2937'),
            leftIndent=20
        )
        
        elements = []
        
        # Title
        elements.append(Paragraph(f"Analysis Report", title_style))
        elements.append(Spacer(1, 12))
        
        # File info table
        file_info = [
            ["File:", session['file_name']],
            ["Rows:", str(session['total_rows'])],
            ["Columns:", ", ".join(session['columns'][:8]) + ("..." if len(session['columns']) > 8 else "")],
            ["Date:", session['created_at'][:19].replace('T', ' ')]
        ]
        
        info_table = PDFTable(file_info, colWidths=[1.2*inch, 4.5*inch])
        info_table.setStyle(TableStyle([
            ('FONTNAME', (0, 0), (0, -1), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, -1), 10),
            ('TEXTCOLOR', (0, 0), (0, -1), colors.HexColor('#6B7280')),
            ('VALIGN', (0, 0), (-1, -1), 'TOP'),
            ('BOTTOMPADDING', (0, 0), (-1, -1), 8),
        ]))
        elements.append(info_table)
        elements.append(Spacer(1, 20))
        
        # Q&A Section
        if session.get("messages"):
            elements.append(Paragraph("Questions & Answers", heading_style))
            elements.append(Spacer(1, 10))
            
            for i, msg in enumerate(session["messages"], 1):
                # Question
                q_text = f"<b>Q{i}:</b> {msg['question']}"
                elements.append(Paragraph(q_text, question_style))
                
                # Answer - handle line breaks
                answer_text = msg['answer'].replace('\n', '<br/>')
                a_text = f"<font color='#22C55E'><b>A{i}:</b></font> {answer_text}"
                elements.append(Paragraph(a_text, answer_style))
                elements.append(Spacer(1, 10))
        else:
            elements.append(Paragraph("No questions asked yet.", styles['Normal']))
        
        # Footer
        elements.append(Spacer(1, 30))
        footer_style = ParagraphStyle(
            'Footer',
            parent=styles['Normal'],
            fontSize=9,
            textColor=colors.HexColor('#9CA3AF')
        )
        elements.append(Paragraph("Generated by Planet Knowledge - Excel/CSV Analyzer with Gemini AI", footer_style))
        
        # Build PDF
        doc.build(elements)
        buffer.seek(0)
        
        filename = f"analysis_{session['file_name'].rsplit('.', 1)[0]}_{session_id[:8]}.pdf"
        
        return StreamingResponse(
            buffer,
            media_type="application/pdf",
            headers={"Content-Disposition": f"attachment; filename={filename}"}
        )
    
    return router
