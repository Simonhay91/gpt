"""Temporary file upload for chat — one-shot files used by AI, not saved to sources"""
import uuid
import base64
import logging
from pathlib import Path
from datetime import datetime, timezone
from io import BytesIO

from fastapi import APIRouter, HTTPException, Depends, UploadFile, File, Form
from pydantic import BaseModel

from middleware.auth import get_current_user
from db.connection import get_db
from services.rag import get_embedding
from services.file_processor import (
    extract_text_from_pdf,
    extract_text_from_docx,
    extract_text_from_xlsx,
    extract_text_from_csv,
    extract_text_from_image,
    chunk_text,
    chunk_tabular_text,
)

logger = logging.getLogger(__name__)
router = APIRouter(prefix="/api", tags=["temp-files"])

TEMP_DIR = Path("/tmp/planet_temp_files")
TEMP_DIR.mkdir(exist_ok=True)

UPLOAD_DIR = Path(__file__).parent.parent / "uploads"
UPLOAD_DIR.mkdir(exist_ok=True)

MAX_TEMP_SIZE = 20 * 1024 * 1024  # 20 MB

# mime → file_type used internally
MIME_TO_TYPE = {
    "image/jpeg": "image",
    "image/jpg": "image",
    "image/png": "image",
    "application/pdf": "pdf",
    "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet": "xlsx",
    "application/vnd.ms-excel": "xlsx",
    "text/csv": "csv",
    "application/csv": "csv",
    "application/vnd.openxmlformats-officedocument.wordprocessingml.document": "docx",
}

EXT_TO_TYPE = {
    "jpg": "image", "jpeg": "image", "png": "image",
    "pdf": "pdf",
    "xlsx": "xlsx", "xls": "xlsx",
    "csv": "csv",
    "docx": "docx",
}

TYPE_TO_MIME = {
    "image": "image/png",
    "pdf": "application/pdf",
    "xlsx": "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    "csv": "text/csv",
    "docx": "application/vnd.openxmlformats-officedocument.wordprocessingml.document",
}


def _resolve_type(mime: str, filename: str) -> str:
    """Determine internal file_type from mime or filename extension."""
    if mime in MIME_TO_TYPE:
        return MIME_TO_TYPE[mime]
    ext = (filename or "").rsplit(".", 1)[-1].lower()
    if ext in EXT_TO_TYPE:
        return EXT_TO_TYPE[ext]
    raise ValueError(f"Unsupported type: mime={mime}, ext={ext}")


# ==================== UPLOAD TEMP FILE ====================

@router.post("/chat/upload-temp")
async def upload_temp_file(
    file: UploadFile = File(...),
    chat_id: str = Form(...),
    current_user: dict = Depends(get_current_user),
):
    """Upload a temporary file for one-shot AI chat context. Not persisted to sources."""
    content = await file.read()

    if len(content) > MAX_TEMP_SIZE:
        raise HTTPException(status_code=400, detail="File exceeds 20 MB limit")

    try:
        file_type = _resolve_type(file.content_type or "", file.filename or "")
    except ValueError:
        raise HTTPException(
            status_code=400,
            detail="Unsupported file type. Allowed: JPG, PNG, PDF, XLSX, XLS, CSV, DOCX",
        )

    temp_file_id = str(uuid.uuid4())
    safe_name = (file.filename or "file").replace("/", "_").replace("..", "_")
    stored_path = TEMP_DIR / f"{temp_file_id}_{safe_name}"
    stored_path.write_bytes(content)

    # Extract content for preview
    content_text = ""
    try:
        if file_type == "image":
            content_text = "[Изображение прикреплено]"
        elif file_type == "pdf":
            # Use unified PDF extractor with OCR fallback for scanned PDFs
            content_text = extract_text_from_pdf(content)
        elif file_type in ("xlsx", "xls"):
            content_text = extract_text_from_xlsx(content)
        elif file_type == "csv":
            content_text = extract_text_from_csv(content)
        elif file_type == "docx":
            content_text = extract_text_from_docx(content)
    except Exception as e:
        logger.error(f"Temp file extraction error ({file_type}): {e}")
        content_text = ""

    # Persist extracted content to chat.tempFiles so subsequent messages can use it
    if chat_id and content_text:
        db = get_db()
        await db.chats.update_one(
            {"id": chat_id},
            {
                "$push": {
                    "tempFiles": {
                        "id": temp_file_id,
                        "filename": file.filename,
                        "fileType": file_type,
                        "content": content_text[:12000],
                        "uploadedAt": datetime.now(timezone.utc).isoformat(),
                    }
                }
            }
        )

    return {
        "temp_file_id": temp_file_id,
        "filename": file.filename,
        "file_type": file_type,
        "content_preview": content_text[:500] if content_text else "",
        "stored_path": str(stored_path),
    }


# ==================== SAVE TEMP FILE TO SOURCES ====================

class SaveTempSourceRequest(BaseModel):
    temp_file_id: str
    filename: str
    file_type: str
    chat_id: str
    project_id: str


@router.post("/chat/save-temp-to-source")
async def save_temp_to_source(
    data: SaveTempSourceRequest,
    current_user: dict = Depends(get_current_user),
):
    """Promote a temporary file to a permanent project source."""
    db = get_db()

    # Find the temp file
    matches = list(TEMP_DIR.glob(f"{data.temp_file_id}_*"))
    if not matches:
        raise HTTPException(status_code=404, detail="Temporary file not found or already deleted")

    temp_path = matches[0]
    content = temp_path.read_bytes()

    # Extract text for chunking
    if data.file_type == "pdf":
        # Unified PDF extractor includes OCR fallback for scanned PDFs
        extracted_text = extract_text_from_pdf(content)
    elif data.file_type in ("xlsx", "xls"):
        extracted_text = extract_text_from_xlsx(content)
    elif data.file_type == "csv":
        extracted_text = extract_text_from_csv(content)
    elif data.file_type == "docx":
        extracted_text = extract_text_from_docx(content)
    elif data.file_type == "image":
        extracted_text = extract_text_from_image(content)
    else:
        extracted_text = ""

    # Determine file extension for storage
    ext = data.filename.rsplit(".", 1)[-1].lower() if "." in data.filename else data.file_type
    source_id = str(uuid.uuid4())
    storage_filename = f"{source_id}.{ext}"
    storage_path = UPLOAD_DIR / storage_filename
    storage_path.write_bytes(content)

    # Chunk
    if data.file_type in ("xlsx", "xls", "csv"):
        chunks = chunk_tabular_text(extracted_text)
    else:
        chunks = chunk_text(extracted_text) if extracted_text else []

    # Extract sheet names for xlsx
    sheet_names = []
    if data.file_type in ("xlsx", "xls"):
        try:
            from openpyxl import load_workbook
            wb = load_workbook(BytesIO(content), read_only=True, data_only=True)
            sheet_names = wb.sheetnames
            wb.close()
        except Exception:
            pass

    mime = TYPE_TO_MIME.get(data.file_type, "application/octet-stream")

    source_doc = {
        "id": source_id,
        "projectId": data.project_id,
        "kind": "file",
        "originalName": data.filename,
        "url": None,
        "mimeType": mime,
        "sizeBytes": len(content),
        "storagePath": storage_filename,
        "sheetNames": sheet_names,
        "ownerId": current_user["id"],
        "level": "project",
        "status": "active",
        "createdAt": datetime.now(timezone.utc).isoformat(),
    }
    await db.sources.insert_one(source_doc)

    for i, chunk_content in enumerate(chunks):
        embedding = await get_embedding(chunk_content)
        await db.source_chunks.insert_one({
            "id": str(uuid.uuid4()),
            "sourceId": source_id,
            "projectId": data.project_id,
            "chunkIndex": i,
            "content": chunk_content,
            "embedding": embedding,
            "createdAt": datetime.now(timezone.utc).isoformat(),
        })

    # Clean up temp file
    try:
        temp_path.unlink()
    except Exception:
        pass

    return {
        "source_id": source_id,
        "chunk_count": len(chunks),
        "message": f"'{data.filename}' сохранён в источники проекта",
    }
