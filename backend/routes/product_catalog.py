"""Product Catalog routes"""
from fastapi import APIRouter, HTTPException, Depends, UploadFile, File, Query, Form
from fastapi.responses import StreamingResponse
from typing import List, Optional
from datetime import datetime, timezone
from uuid import uuid4
import logging
import csv
import io
import os
import re
import json

from models.schemas import (
    ProductCatalogCreate,
    ProductCatalogUpdate,
    ProductCatalogResponse,
    ProductRelationCreate,
    ProductImportResult,
    ProductMatchRequest,
    ProductMatchResult
)
from middleware.auth import get_current_user, is_admin
from db.connection import get_db

logger = logging.getLogger(__name__)
router = APIRouter(prefix="/api", tags=["product_catalog"])


# ==================== HELPER FUNCTIONS ====================

def now_iso():
    return datetime.now(timezone.utc).isoformat()


def can_edit_catalog(current_user: dict) -> bool:
    """Check if user can edit product catalog (Admin or Manager)"""
    if is_admin(current_user.get("email", "")):
        return True
    # Check if user is manager of any department
    return current_user.get("isManager", False)


async def check_manager_status(current_user: dict) -> bool:
    """Check if user is manager of any department"""
    db = get_db()
    user_id = current_user["id"]
    
    # Check if admin
    if is_admin(current_user.get("email", "")):
        return True
    
    # Check if manager of any department
    manager_dept = await db.departments.find_one(
        {"managers": user_id},
        {"_id": 0, "id": 1}
    )
    return manager_dept is not None


# ==================== CRUD ENDPOINTS ====================

@router.get("/product-catalog", response_model=List[ProductCatalogResponse])
async def list_products(
    search: Optional[str] = Query(None, description="Search in title, article_number, vendor, aliases"),
    root_category: Optional[str] = Query(None),
    lvl1_subcategory: Optional[str] = Query(None),
    vendor: Optional[str] = Query(None),
    is_active: Optional[bool] = Query(None),
    limit: int = Query(100, le=1000),
    offset: int = Query(0),
    current_user: dict = Depends(get_current_user)
):
    """Get all products with search and filters"""
    db = get_db()
    
    query = {}
    
    # Search across multiple fields
    if search:
        search_regex = {"$regex": search, "$options": "i"}
        query["$or"] = [
            {"title_en": search_regex},
            {"article_number": search_regex},
            {"vendor": search_regex},
            {"aliases": search_regex},
            {"product_model": search_regex},
            {"crm_code": search_regex}
        ]
    
    # Filters
    if root_category:
        query["root_category"] = root_category
    if lvl1_subcategory:
        query["lvl1_subcategory"] = lvl1_subcategory
    if vendor:
        query["vendor"] = vendor
    if is_active is not None:
        query["is_active"] = is_active
    
    products = await db.product_catalog.find(
        query, {"_id": 0}
    ).sort("title_en", 1).skip(offset).limit(limit).to_list(limit)
    
    return products


@router.get("/product-catalog/categories")
async def get_categories(current_user: dict = Depends(get_current_user)):
    """Get unique categories and vendors for filters"""
    db = get_db()
    
    pipeline = [
        {"$match": {"is_active": True}},
        {"$group": {
            "_id": None,
            "root_categories": {"$addToSet": "$root_category"},
            "lvl1_subcategories": {"$addToSet": "$lvl1_subcategory"},
            "vendors": {"$addToSet": "$vendor"}
        }}
    ]
    
    result = await db.product_catalog.aggregate(pipeline).to_list(1)
    
    if result:
        data = result[0]
        return {
            "root_categories": sorted([c for c in data.get("root_categories", []) if c]),
            "lvl1_subcategories": sorted([c for c in data.get("lvl1_subcategories", []) if c]),
            "vendors": sorted([v for v in data.get("vendors", []) if v])
        }
    
    return {"root_categories": [], "lvl1_subcategories": [], "vendors": []}


@router.get("/product-catalog/stats")
async def get_stats(current_user: dict = Depends(get_current_user)):
    """Get product catalog statistics"""
    db = get_db()
    
    total = await db.product_catalog.count_documents({})
    active = await db.product_catalog.count_documents({"is_active": True})
    inactive = await db.product_catalog.count_documents({"is_active": False})
    
    # Get last sync date
    last_synced = await db.product_catalog.find_one(
        {"last_synced_at": {"$ne": None}},
        {"_id": 0, "last_synced_at": 1},
        sort=[("last_synced_at", -1)]
    )
    
    return {
        "total": total,
        "active": active,
        "inactive": inactive,
        "last_synced_at": last_synced.get("last_synced_at") if last_synced else None
    }


@router.get("/product-catalog/{product_id}", response_model=ProductCatalogResponse)
async def get_product(product_id: str, current_user: dict = Depends(get_current_user)):
    """Get single product with relations"""
    db = get_db()
    
    product = await db.product_catalog.find_one({"id": product_id}, {"_id": 0})
    if not product:
        raise HTTPException(status_code=404, detail="Product not found")
    
    return product


@router.post("/product-catalog", response_model=ProductCatalogResponse)
async def create_product(data: ProductCatalogCreate, current_user: dict = Depends(get_current_user)):
    """Create a new product (Admin/Manager only)"""
    db = get_db()
    
    # Check permission
    has_permission = await check_manager_status(current_user)
    if not has_permission:
        raise HTTPException(status_code=403, detail="Only Admin or Manager can create products")
    
    # Check for duplicate article_number
    existing = await db.product_catalog.find_one(
        {"article_number": data.article_number},
        {"_id": 0, "id": 1}
    )
    if existing:
        raise HTTPException(status_code=400, detail=f"Product with article number '{data.article_number}' already exists")
    
    now = now_iso()
    product = {
        "id": str(uuid4()),
        "article_number": data.article_number,
        "title_en": data.title_en,
        "crm_code": data.crm_code,
        "root_category": data.root_category,
        "lvl1_subcategory": data.lvl1_subcategory,
        "lvl2_subcategory": data.lvl2_subcategory,
        "lvl3_subcategory": data.lvl3_subcategory,
        "vendor": data.vendor,
        "description": data.description,
        "features": data.features,
        "attribute_values": data.attribute_values,
        "product_model": data.product_model,
        "datasheet_url": data.datasheet_url,
        "aliases": data.aliases or [],
        "price": data.price,
        "relations": [],
        "extra_fields": data.extra_fields,
        "is_active": True,
        "last_synced_at": None,
        "source": "manual",
        "created_by": current_user["id"],
        "updated_by": None,
        "created_at": now,
        "updated_at": None
    }
    
    await db.product_catalog.insert_one(product)
    
    return ProductCatalogResponse(**product)


@router.put("/product-catalog/{product_id}", response_model=ProductCatalogResponse)
async def update_product(
    product_id: str,
    data: ProductCatalogUpdate,
    current_user: dict = Depends(get_current_user)
):
    """Update a product (Admin/Manager only)"""
    db = get_db()
    
    # Check permission
    has_permission = await check_manager_status(current_user)
    if not has_permission:
        raise HTTPException(status_code=403, detail="Only Admin or Manager can update products")
    
    product = await db.product_catalog.find_one({"id": product_id}, {"_id": 0})
    if not product:
        raise HTTPException(status_code=404, detail="Product not found")
    
    update_data = {"updated_by": current_user["id"], "updated_at": now_iso()}
    
    for field, value in data.dict(exclude_unset=True).items():
        if value is not None:
            update_data[field] = value
    
    await db.product_catalog.update_one(
        {"id": product_id},
        {"$set": update_data}
    )
    
    updated_product = await db.product_catalog.find_one({"id": product_id}, {"_id": 0})
    return ProductCatalogResponse(**updated_product)


@router.delete("/product-catalog/{product_id}")
async def delete_product(product_id: str, current_user: dict = Depends(get_current_user)):
    """Soft delete a product (Admin/Manager only)"""
    db = get_db()
    
    # Check permission
    has_permission = await check_manager_status(current_user)
    if not has_permission:
        raise HTTPException(status_code=403, detail="Only Admin or Manager can delete products")
    
    product = await db.product_catalog.find_one({"id": product_id}, {"_id": 0})
    if not product:
        raise HTTPException(status_code=404, detail="Product not found")
    
    await db.product_catalog.update_one(
        {"id": product_id},
        {"$set": {"is_active": False, "updated_by": current_user["id"], "updated_at": now_iso()}}
    )
    
    return {"success": True, "message": "Product deactivated"}


# ==================== RELATIONS ENDPOINTS ====================

@router.post("/product-catalog/{product_id}/relations")
async def add_relation(
    product_id: str,
    data: ProductRelationCreate,
    current_user: dict = Depends(get_current_user)
):
    """Add a relation to a product (Admin/Manager only)"""
    db = get_db()
    
    # Check permission
    has_permission = await check_manager_status(current_user)
    if not has_permission:
        raise HTTPException(status_code=403, detail="Only Admin or Manager can manage relations")
    
    # Check both products exist
    product = await db.product_catalog.find_one({"id": product_id}, {"_id": 0})
    if not product:
        raise HTTPException(status_code=404, detail="Product not found")
    
    related_product = await db.product_catalog.find_one({"id": data.product_id}, {"_id": 0})
    if not related_product:
        raise HTTPException(status_code=404, detail="Related product not found")
    
    if product_id == data.product_id:
        raise HTTPException(status_code=400, detail="Cannot create relation to self")
    
    # Check if relation already exists
    existing_relations = product.get("relations", [])
    for rel in existing_relations:
        if rel["product_id"] == data.product_id:
            raise HTTPException(status_code=400, detail="Relation already exists")
    
    # Add relation
    relation = {"product_id": data.product_id, "relation_type": data.relation_type}
    
    await db.product_catalog.update_one(
        {"id": product_id},
        {
            "$push": {"relations": relation},
            "$set": {"updated_by": current_user["id"], "updated_at": now_iso()}
        }
    )
    
    # Add reverse relation for bidirectional types
    if data.relation_type in ["compatible", "bundle"]:
        reverse_relation = {"product_id": product_id, "relation_type": data.relation_type}
        await db.product_catalog.update_one(
            {"id": data.product_id},
            {
                "$push": {"relations": reverse_relation},
                "$set": {"updated_by": current_user["id"], "updated_at": now_iso()}
            }
        )
    
    return {"success": True, "message": "Relation added"}


@router.delete("/product-catalog/{product_id}/relations/{related_id}")
async def remove_relation(
    product_id: str,
    related_id: str,
    current_user: dict = Depends(get_current_user)
):
    """Remove a relation from a product (Admin/Manager only)"""
    db = get_db()
    
    # Check permission
    has_permission = await check_manager_status(current_user)
    if not has_permission:
        raise HTTPException(status_code=403, detail="Only Admin or Manager can manage relations")
    
    # Remove from main product
    result = await db.product_catalog.update_one(
        {"id": product_id},
        {
            "$pull": {"relations": {"product_id": related_id}},
            "$set": {"updated_by": current_user["id"], "updated_at": now_iso()}
        }
    )
    
    if result.modified_count == 0:
        raise HTTPException(status_code=404, detail="Relation not found")
    
    # Remove reverse relation
    await db.product_catalog.update_one(
        {"id": related_id},
        {
            "$pull": {"relations": {"product_id": product_id}},
            "$set": {"updated_by": current_user["id"], "updated_at": now_iso()}
        }
    )
    
    return {"success": True, "message": "Relation removed"}


# ==================== IMPORT ENDPOINTS ====================

# Column mappings for CSV import
CSV_COLUMN_MAP = {
    "article number": "article_number",
    "title en": "title_en",
    "crm code": "crm_code",
    "root category": "root_category",
    "lvl 1 subcategory": "lvl1_subcategory",
    "lvl1 subcategory": "lvl1_subcategory",
    "lvl 2 subcategory": "lvl2_subcategory",
    "lvl2 subcategory": "lvl2_subcategory",
    "lvl 3 subcategory": "lvl3_subcategory",
    "lvl3 subcategory": "lvl3_subcategory",
    "vendor": "vendor",
    "description": "description",
    "features": "features",
    "attribute values": "attribute_values",
    "product model": "product_model",
    "datasheet url": "datasheet_url",
    "price": "price"
}

KNOWN_COLUMNS = set(CSV_COLUMN_MAP.keys())


@router.post("/product-catalog/import/preview")
async def preview_import(
    file: UploadFile = File(...),
    current_user: dict = Depends(get_current_user)
):
    """Preview CSV import - detect columns and show unknown ones"""
    db = get_db()
    
    # Check permission
    has_permission = await check_manager_status(current_user)
    if not has_permission:
        raise HTTPException(status_code=403, detail="Only Admin or Manager can import products")
    
    # Read file
    content = await file.read()
    try:
        text = content.decode('utf-8')
    except UnicodeDecodeError:
        text = content.decode('cp1251')  # Try Windows encoding
    
    reader = csv.DictReader(io.StringIO(text))
    headers = [h.strip().lower() for h in reader.fieldnames] if reader.fieldnames else []
    
    # Detect known and unknown columns
    known = []
    unknown = []
    
    for h in headers:
        if h in KNOWN_COLUMNS:
            known.append({"original": h, "mapped_to": CSV_COLUMN_MAP[h]})
        else:
            unknown.append(h)
    
    # Count rows
    rows = list(reader)
    
    return {
        "filename": file.filename,
        "total_rows": len(rows),
        "known_columns": known,
        "unknown_columns": unknown,
        "sample_data": rows[:5] if rows else []
    }


@router.post("/product-catalog/import", response_model=ProductImportResult)
async def import_products(
    file: UploadFile = File(...),
    extra_columns: Optional[str] = Query(None, description="Comma-separated list of extra columns to import"),
    current_user: dict = Depends(get_current_user)
):
    """Import products from CSV"""
    db = get_db()
    
    # Check permission
    has_permission = await check_manager_status(current_user)
    if not has_permission:
        raise HTTPException(status_code=403, detail="Only Admin or Manager can import products")
    
    # Parse extra columns
    extra_cols = [c.strip().lower() for c in extra_columns.split(",")] if extra_columns else []
    
    # Read file
    content = await file.read()
    try:
        text = content.decode('utf-8')
    except UnicodeDecodeError:
        text = content.decode('cp1251')
    
    reader = csv.DictReader(io.StringIO(text))
    rows = list(reader)
    
    now = now_iso()
    added = 0
    updated = 0
    skipped = 0
    errors = []
    
    # Track article numbers in this import
    imported_articles = set()
    
    for i, row in enumerate(rows):
        # Normalize keys
        row_normalized = {k.strip().lower(): v.strip() if v else None for k, v in row.items()}
        
        # Get required fields
        article_number = row_normalized.get("article number")
        title_en = row_normalized.get("title en")
        
        if not article_number or not title_en:
            errors.append(f"Row {i+2}: Missing required field (article_number or title_en)")
            skipped += 1
            continue
        
        imported_articles.add(article_number)
        
        # Build product data
        product_data = {
            "article_number": article_number,
            "title_en": title_en,
            "updated_by": current_user["id"],
            "updated_at": now,
            "last_synced_at": now,
            "source": "csv_import",
            "is_active": True
        }
        
        # Map known columns
        for csv_col, db_col in CSV_COLUMN_MAP.items():
            if csv_col in row_normalized and row_normalized[csv_col]:
                value = row_normalized[csv_col]
                if db_col == "price":
                    try:
                        value = float(value.replace(",", ".").replace(" ", ""))
                    except:
                        value = None
                product_data[db_col] = value
        
        # Map extra columns
        if extra_cols:
            extra_fields = {}
            for col in extra_cols:
                if col in row_normalized and row_normalized[col]:
                    extra_fields[col] = row_normalized[col]
            if extra_fields:
                product_data["extra_fields"] = extra_fields
        
        # Check if exists
        existing = await db.product_catalog.find_one(
            {"article_number": article_number},
            {"_id": 0, "id": 1, "relations": 1, "aliases": 1, "extra_fields": 1}
        )
        
        if existing:
            # Update existing - preserve relations, aliases, extra_fields
            preserve_fields = {}
            if existing.get("relations"):
                preserve_fields["relations"] = existing["relations"]
            if existing.get("aliases"):
                preserve_fields["aliases"] = existing["aliases"]
            if existing.get("extra_fields") and product_data.get("extra_fields"):
                # Merge extra fields
                merged_extra = {**existing["extra_fields"], **product_data["extra_fields"]}
                product_data["extra_fields"] = merged_extra
            elif existing.get("extra_fields"):
                preserve_fields["extra_fields"] = existing["extra_fields"]
            
            await db.product_catalog.update_one(
                {"id": existing["id"]},
                {"$set": {**product_data, **preserve_fields}}
            )
            updated += 1
        else:
            # Create new
            product_data["id"] = str(uuid4())
            product_data["relations"] = []
            product_data["aliases"] = []
            product_data["created_by"] = current_user["id"]
            product_data["created_at"] = now
            
            await db.product_catalog.insert_one(product_data)
            added += 1
    
    # Deactivate products not in this import
    deactivated_result = await db.product_catalog.update_many(
        {
            "article_number": {"$nin": list(imported_articles)},
            "is_active": True,
            "source": "csv_import"
        },
        {"$set": {"is_active": False, "updated_by": current_user["id"], "updated_at": now}}
    )
    deactivated = deactivated_result.modified_count
    
    logger.info(f"Product import completed: added={added}, updated={updated}, deactivated={deactivated}, skipped={skipped}")
    
    return ProductImportResult(
        added=added,
        updated=updated,
        deactivated=deactivated,
        skipped=skipped,
        errors=errors[:10]  # Limit errors in response
    )


# ==================== TENDER MATCHING ENDPOINT ====================

@router.post("/product-catalog/match", response_model=List[ProductMatchResult])
async def match_products(
    data: ProductMatchRequest,
    current_user: dict = Depends(get_current_user)
):
    """Match product titles against catalog (for tender analysis)"""
    db = get_db()
    
    results = []
    
    for title in data.titles:
        if not title or not title.strip():
            continue
        
        title_clean = title.strip().lower()
        
        # Search by exact article number first
        product = await db.product_catalog.find_one(
            {"article_number": {"$regex": f"^{title_clean}$", "$options": "i"}, "is_active": True},
            {"_id": 0}
        )
        
        if product:
            results.append(ProductMatchResult(query=title, matched=product, confidence=1.0))
            continue
        
        # Search by title_en, product_model, aliases
        product = await db.product_catalog.find_one(
            {
                "$or": [
                    {"title_en": {"$regex": title_clean, "$options": "i"}},
                    {"product_model": {"$regex": title_clean, "$options": "i"}},
                    {"aliases": {"$regex": title_clean, "$options": "i"}}
                ],
                "is_active": True
            },
            {"_id": 0}
        )
        
        if product:
            results.append(ProductMatchResult(query=title, matched=product, confidence=0.8))
        else:
            results.append(ProductMatchResult(query=title, matched=None, confidence=0.0))
    
    return results


# ==================== FILE MATCHING ENDPOINT ====================

def _parse_names_from_excel(content: bytes, name_column: str) -> List[str]:
    """Parse product names from Excel/CSV file."""
    import pandas as pd
    try:
        df = pd.read_excel(io.BytesIO(content), header=0)
    except Exception:
        try:
            text = content.decode("utf-8")
        except UnicodeDecodeError:
            text = content.decode("cp1251")
        df = pd.read_csv(io.StringIO(text))

    if name_column.isdigit():
        col_idx = int(name_column)
        if col_idx < len(df.columns):
            col = df.columns[col_idx]
        else:
            col = df.columns[0]
    else:
        # Try to find column by name (case-insensitive)
        matched = [c for c in df.columns if name_column.lower() in str(c).lower()]
        col = matched[0] if matched else df.columns[0]

    names = [str(v).strip() for v in df[col].dropna() if str(v).strip() and str(v).strip().lower() != "nan"]
    return names


def _parse_names_from_docx(content: bytes) -> List[str]:
    from docx import Document
    doc = Document(io.BytesIO(content))
    names = []
    for para in doc.paragraphs:
        line = para.text.strip()
        if line:
            names.append(line)
    return names


def _parse_names_from_pdf(content: bytes) -> List[str]:
    import pdfplumber
    names = []
    with pdfplumber.open(io.BytesIO(content)) as pdf:
        for page in pdf.pages:
            text = page.extract_text() or ""
            for line in text.splitlines():
                line = line.strip()
                if line:
                    names.append(line)
    return names


async def _search_candidates(db, query: str, limit: int = 6) -> List[dict]:
    """Fast text search for candidate products."""
    words = [w for w in re.findall(r'\w+', query) if len(w) > 2]
    if not words:
        return []
    regex_parts = [{"title_en": {"$regex": w, "$options": "i"}} for w in words[:5]]
    regex_parts += [{"article_number": {"$regex": w, "$options": "i"}} for w in words[:3]]
    regex_parts += [{"product_model": {"$regex": w, "$options": "i"}} for w in words[:3]]
    regex_parts += [{"aliases": {"$elemMatch": {"$regex": w, "$options": "i"}}} for w in words[:3]]
    results = await db.product_catalog.find(
        {"is_active": True, "$or": regex_parts},
        {"_id": 0, "id": 1, "title_en": 1, "article_number": 1, "crm_code": 1,
         "datasheet_url": 1, "vendor": 1, "product_model": 1, "description": 1}
    ).limit(limit).to_list(limit)
    return results


async def _claude_match_batch(batch: List[dict], mode: str) -> List[dict]:
    """
    Ask Claude to match a batch of {customer_name, candidates[]} items.
    Returns list of match results.
    """
    import anthropic
    CLAUDE_API_KEY = os.environ.get("CLAUDE_API_KEY", "")
    client = anthropic.Anthropic(api_key=CLAUDE_API_KEY)

    code_field = "article_number" if mode == "article" else "crm_code"

    batch_text = ""
    for i, item in enumerate(batch):
        batch_text += f"\n--- Product {i+1} ---\n"
        batch_text += f"Customer name: {item['customer_name']}\n"
        if item["candidates"]:
            batch_text += "Catalog candidates:\n"
            for j, c in enumerate(item["candidates"]):
                batch_text += (
                    f"  [{j+1}] title={c.get('title_en','')} | "
                    f"article={c.get('article_number','')} | "
                    f"crm={c.get('crm_code','')} | "
                    f"model={c.get('product_model','')} | "
                    f"vendor={c.get('vendor','')}\n"
                )
        else:
            batch_text += "Catalog candidates: (none found)\n"

    prompt = f"""You are a product matching assistant for a fiber optics / network equipment catalog.

For each product below, find the best match from the given catalog candidates.
Return a JSON array with one object per product in the same order:
{{
  "matched_title": "<catalog title_en or empty string>",
  "matched_article": "<article_number or empty string>",
  "matched_crm": "<crm_code or empty string>",
  "matched_datasheet": "<datasheet_url or empty string>",
  "confidence": <integer 0-100>,
  "comment": "<short explanation in English, max 15 words>"
}}

Rules:
- If no good match exists, set confidence=0 and leave matched fields empty.
- confidence=90+ means near-identical, 70-89 means close match (minor spec diff), 50-69 means probable match, <50 means uncertain.
- Keep comment concise.

Products to match:
{batch_text}

Return ONLY the JSON array, no extra text."""

    response = client.messages.create(
        model="claude-sonnet-4-20250514",
        max_tokens=2000,
        messages=[{"role": "user", "content": prompt}]
    )
    raw = response.content[0].text.strip()
    # Extract JSON array
    match = re.search(r'\[.*\]', raw, re.DOTALL)
    if match:
        return json.loads(match.group(0))
    return [{"matched_title": "", "matched_article": "", "matched_crm": "",
             "matched_datasheet": "", "confidence": 0, "comment": "Parse error"} for _ in batch]


def _build_excel(rows: List[dict]) -> bytes:
    """Build result Excel file in memory."""
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Matched Products"

    headers = [
        "Customer Product Name",
        "Matched Catalog Name",
        "Article Number",
        "CRM Code",
        "Datasheet",
        "Confidence %",
        "AI Comment"
    ]
    header_fill = PatternFill("solid", fgColor="1E3A5F")
    header_font = Font(bold=True, color="FFFFFF")

    for col_idx, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_idx, value=h)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")

    for row_idx, r in enumerate(rows, 2):
        conf = r.get("confidence", 0)
        row_fill = None
        if conf >= 85:
            row_fill = PatternFill("solid", fgColor="D6EFCD")  # green
        elif conf >= 60:
            row_fill = PatternFill("solid", fgColor="FFF2CC")  # yellow
        elif conf > 0:
            row_fill = PatternFill("solid", fgColor="FCE4D6")  # orange

        values = [
            r.get("customer_name", ""),
            r.get("matched_title", ""),
            r.get("matched_article", ""),
            r.get("matched_crm", ""),
            r.get("matched_datasheet", ""),
            conf,
            r.get("comment", "")
        ]
        for col_idx, val in enumerate(values, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=val)
            if row_fill:
                cell.fill = row_fill

    # Column widths
    widths = [40, 45, 18, 18, 40, 14, 40]
    for col_idx, w in enumerate(widths, 1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(col_idx)].width = w

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.read()


@router.post("/product-catalog/match-file")
async def match_file(
    file: UploadFile = File(...),
    mode: str = Form("article"),          # "article" or "crm"
    name_column: str = Form("0"),          # column index or name
    current_user: dict = Depends(get_current_user)
):
    """
    Upload a customer file (Excel/CSV/DOCX/PDF) with product names.
    Returns an Excel file with matched catalog products, confidence scores, and AI comments.
    """
    db = get_db()
    content = await file.read()
    filename = (file.filename or "").lower()

    # 1. Parse product names from file
    try:
        if filename.endswith(".pdf"):
            customer_names = _parse_names_from_pdf(content)
        elif filename.endswith(".docx"):
            customer_names = _parse_names_from_docx(content)
        else:
            customer_names = _parse_names_from_excel(content, name_column)
    except Exception as e:
        raise HTTPException(status_code=400, detail=f"Failed to parse file: {e}")

    if not customer_names:
        raise HTTPException(status_code=400, detail="No product names found in file")

    # Limit to 200 rows per request
    customer_names = customer_names[:200]

    # 2. For each name, find text-based candidates from catalog
    batch_items = []
    for name in customer_names:
        candidates = await _search_candidates(db, name, limit=6)
        batch_items.append({"customer_name": name, "candidates": candidates})

    # 3. Claude batch matching (10 products per call)
    results = []
    BATCH_SIZE = 10
    for i in range(0, len(batch_items), BATCH_SIZE):
        batch = batch_items[i:i + BATCH_SIZE]
        try:
            matches = await _claude_match_batch(batch, mode)
        except Exception as e:
            logger.error(f"Claude batch matching error: {e}")
            matches = [{"matched_title": "", "matched_article": "", "matched_crm": "",
                        "matched_datasheet": "", "confidence": 0, "comment": "Error"} for _ in batch]
        for item, match in zip(batch, matches):
            results.append({
                "customer_name": item["customer_name"],
                "matched_title": match.get("matched_title", ""),
                "matched_article": match.get("matched_article", ""),
                "matched_crm": match.get("matched_crm", ""),
                "matched_datasheet": match.get("matched_datasheet", ""),
                "confidence": match.get("confidence", 0),
                "comment": match.get("comment", "")
            })

    # 4. Build and return Excel
    excel_bytes = _build_excel(results)
    return StreamingResponse(
        io.BytesIO(excel_bytes),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=matched_products.xlsx"}
    )
