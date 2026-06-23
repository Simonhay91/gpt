"""Excel file generation service"""
import json
import math
import uuid
import logging
from datetime import datetime, timezone
from pathlib import Path
from typing import Optional, Tuple

import pandas as pd

logger = logging.getLogger(__name__)


async def _persist_excel_to_db(db, file_id: str, file_path: str, source_name: str = "result.xlsx"):
    """Mirror the generated Excel into MongoDB so the file survives pod restarts.
    Files are typically <1 MB; well within the 16 MB BSON document limit."""
    try:
        with open(file_path, "rb") as f:
            blob = f.read()
        await db.excel_files.update_one(
            {"id": file_id},
            {"$set": {
                "id": file_id,
                "filename": source_name,
                "mimeType": "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                "size": len(blob),
                "data": blob,
                "createdAt": datetime.now(timezone.utc).isoformat(),
            }},
            upsert=True,
        )
    except Exception as e:
        logger.error(f"persist_excel_to_db failed for {file_id}: {e}")

EXCEL_MIME_TYPES = [
    "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    "application/vnd.ms-excel",
    "text/csv",
    "application/csv",
]

EXCEL_TRIGGER_PHRASES = [
    # ── English — generate ──
    "generate excel", "create excel", "make excel", "create spreadsheet",
    "make spreadsheet", "download excel", "export excel", "export to excel",
    "generate", "download", "export", "get file", "give me file", "save as excel",
    # ── English — edit ──
    "edit", "modify", "update", "change", "fix",
    # ── Russian — generate ──
    "сгенерируй excel", "создай excel", "сделай excel", "скачать excel",
    "сгенерируй таблицу", "создай таблицу", "сделай таблицу",
    "скачай", "скачать", "сохрани", "экспортируй", "генерируй", "создай файл",
    # ── Russian — edit ──
    "редактируй", "измени", "обнови", "исправь", "поменяй",
    # ── Armenian romanized — generate ──
    "excel generacru", "excel sarcru", "excel beri", "excel download ara",
    "avelacru excel", "excel poxi", "kercru excel",
    "generacru", "sarcru", "beri", "tui", "ktur", "bacer", "paterastirel",
    "download ara", "paterastr", "excel baci", "excel kazmir",
    # ── Armenian romanized — edit ──
    "edit ara", "poxi", "kpoxes", "popoxir", "gri", "nerkayacru",
    # ── Armenian unicode — generate ──
    "excel ստեղծիր", "excel բեր", "excel ներբեռնիր",
    "ստեղծիր", "բեր", "տուր", "ներբեռնիր", "պատրաստիր", "գեներացրու",
    # ── Armenian unicode — edit ──
    "փոխիր", "խմբագրիր", "թարմացրու", "ուղղիր", "գրիր",
    # ── Excel-specific ──
    "arajin togh", "առաջին տող", "readme",
]

# When an existing Excel source is found, ALL trigger phrases route to targeted_edit.
# Full generation only happens when there is NO existing file (scratch creation).
EXCEL_EDIT_PHRASES = [
    # ── English ──
    "edit", "modify", "update", "change", "fix",
    "generate", "download", "export", "get file", "give me file",
    # ── Russian ──
    "редактируй", "измени", "обнови", "исправь", "поменяй",
    "скачай", "скачать", "сохрани", "экспортируй", "генерируй",
    # ── Armenian romanized ──
    "edit ara", "poxi", "kpoxes", "popoxir", "gri", "nerkayacru",
    "generacru", "sarcru", "beri", "tui", "ktur", "bacer",
    "download ara", "paterastr", "paterastirel",
    # ── Armenian unicode ──
    "փոխիր", "խմբագրիր", "թարմացրու", "ուղղիր", "գրիր",
    "ստեղծիր", "բեր", "տուր", "ներբեռնիր", "պատրաստիր", "գեներացրու",
    # ── Excel-specific ──
    "arajin togh", "առաջին տող", "readme",
]

# Messages matching these patterns must NEVER trigger Excel edit or generation.
EXCEL_EDIT_SKIP_WORDS = [
    # Sheet info queries — English
    "what sheets", "list sheets", "show sheets", "which sheet", "how many sheets",
    # Sheet info queries — Armenian romanized
    "inch sheeter", "inch sheet", "inch sheter", "sheeter ka", "sheet ka",
    "inch sheeter ka", "qani sheet", "qani sheeter", "sheeter uni", "sheet uni",
    "inch sheeter es tesnum", "inch sheet es tesnum",
    # Sheet info queries — Armenian unicode
    "ինչ sheet", "քանի sheet", "ինչ շիտ", "sheet-եր",
    # Sheet info queries — Russian
    "какие листы", "список листов", "сколько листов", "какие вкладки",
    # General question/info words — Armenian romanized
    "anhaskacox", "inch ka", "inch uni", "inch pes", "vonc",
    "asa indz", "tur indz", "cuyc tur", "cuic tur", "tesnem",
    # Russian question indicators
    "что такое", "что это", "как называется", "расскажи",
    # English info indicators
    "what is", "show me", "tell me", "list all",
]

# Minimum word count for edit trigger — short messages like "generacru" (1 word) ARE valid.
EXCEL_EDIT_MIN_WORDS = 1

UPLOAD_DIR = Path(__file__).parent.parent / "uploads"


def is_excel_trigger(message_content: str) -> bool:
    """Check if message explicitly requests Excel generation or editing."""
    content_lower = message_content.lower()
    # Never trigger if skip words present
    if any(skip in content_lower for skip in EXCEL_EDIT_SKIP_WORDS):
        return False
    return any(phrase in content_lower for phrase in EXCEL_TRIGGER_PHRASES)


def is_edit_trigger(message_content: str) -> bool:
    """Check if message requests targeted cell editing (not full regeneration).

    Guards:
    1. Skip if message contains any EXCEL_EDIT_SKIP_WORDS.
    2. Skip if message is too short (< EXCEL_EDIT_MIN_WORDS words).
    3. Only trigger if an EXCEL_EDIT_PHRASES keyword is present.
    """
    content_lower = message_content.strip().lower()

    # Guard 1: explicit skip words
    if any(skip in content_lower for skip in EXCEL_EDIT_SKIP_WORDS):
        return False

    # Guard 2: too short — "anhaskacox es?" and similar must not edit
    if len(content_lower.split()) < EXCEL_EDIT_MIN_WORDS:
        return False

    # Guard 3: must contain an actual edit keyword
    return any(phrase in content_lower for phrase in EXCEL_EDIT_PHRASES)


def _sanitize_value(v):
    """Coerce a value to a type openpyxl can write to a cell.

    openpyxl accepts: str, int, float, bool, datetime, None.
    Anything else (list, dict, etc.) is stringified so the cell write never crashes.
    """
    if v is None:
        return None
    if isinstance(v, bool):
        return v
    if isinstance(v, (int, float)):
        if math.isnan(v) or math.isinf(v):
            return None
        return v
    if isinstance(v, str):
        return v
    # dicts, lists, or any other type → stringify
    return str(v)


async def targeted_excel_edit(source_file_path: str, instruction: str, claude_client) -> tuple:
    """Edit Excel file — cells, formulas, styles, chart titles, merges, row/col sizes."""
    import openpyxl
    from openpyxl.styles import PatternFill, Font, Alignment
    from openpyxl.utils import get_column_letter

    if not source_file_path.endswith(('.xlsx', '.xlsm', '.xls')):
        return None, None, "Edit is only supported for Excel files (.xlsx). CSV files cannot be edited this way."

    # ── Build full file structure for Claude ──
    _wb_preview = openpyxl.load_workbook(source_file_path, read_only=True, data_only=True)
    file_structure = {}
    for _sheet_name in _wb_preview.sheetnames:
        _ws = _wb_preview[_sheet_name]
        _rows_data = []
        for _row in _ws.iter_rows(values_only=True):
            _rows_data.append(list(_row))
        file_structure[_sheet_name] = {
            "rows": _rows_data,
            "max_row": _ws.max_row,
            "max_col": _ws.max_column,
        }
    _wb_preview.close()

    # Collect chart info from a writable workbook (read_only doesn't expose charts)
    _wb_charts = openpyxl.load_workbook(source_file_path)
    charts_info = {}
    for _sheet_name in _wb_charts.sheetnames:
        _ws_c = _wb_charts[_sheet_name]
        _chart_list = []
        for _i, _chart in enumerate(getattr(_ws_c, '_charts', [])):
            _title = ""
            try:
                _title = str(_chart.title) if _chart.title else ""
            except Exception:
                pass
            _chart_list.append({"index": _i, "title": _title, "type": type(_chart).__name__})
        if _chart_list:
            charts_info[_sheet_name] = _chart_list
    _wb_charts.close()

    if charts_info:
        for _sn, _cl in charts_info.items():
            if _sn in file_structure:
                file_structure[_sn]["charts"] = _cl

    print(f"[EXCEL EDIT DEBUG] Instruction: {instruction}")
    print(f"[EXCEL EDIT DEBUG] File structure: {json.dumps(file_structure, ensure_ascii=False)[:800]}")

    # ── Ask Claude for rich operation list ──
    analysis_response = await claude_client.messages.create(
        model="claude-sonnet-4-6",
        max_tokens=2048,
        system=(
            "You are a full-featured Excel editor. The user's instruction may be in Armenian, Russian, or English.\n"
            "Language hints:\n"
            "- Armenian: poxi/փոխիր=change | gri/գրիր=write | karmir/կարմիր=red | deghin/դեղին=yellow | kanahaguyn/կանաչ=green | spitak/սպիտակ=white | sheganakarguyn/շականակ=brown\n"
            "- Russian: замени=change | напиши=write | красный=red | жёлтый=yellow\n"
            "- English: change/set/write | red/yellow/green\n\n"
            "Return ONLY a JSON array of operations — no markdown, no explanation.\n"
            "Supported operation types:\n"
            '1. {"type":"cell","sheet":"...","cell":"A1","value":"..."}  — set cell value or formula (=SUM(...))\n'
            '2. {"type":"fill","sheet":"...","cell":"A1","color":"FF0000"}  — background color (hex, no #)\n'
            '3. {"type":"font","sheet":"...","cell":"A1","bold":true,"italic":false,"size":14,"color":"FFFFFF"}  — font style\n'
            '4. {"type":"chart_title","sheet":"...","chart_index":0,"title":"New Title"}  — change chart title\n'
            '5. {"type":"chart_fill","sheet":"...","chart_index":0,"color":"FF0000"}  — chart plot area background\n'
            '6. {"type":"merge","sheet":"...","range":"A1:D1"}  — merge cells\n'
            '7. {"type":"unmerge","sheet":"...","range":"A1:D1"}  — unmerge cells\n'
            '8. {"type":"row_height","sheet":"...","row":1,"height":30}  — row height in points\n'
            '9. {"type":"col_width","sheet":"...","col":"A","width":20}  — column width\n'
            "Color names → hex: red=FF0000, yellow=FFFF00, green=00FF00, blue=0000FF, white=FFFFFF, black=000000, orange=FFA500\n"
            "Rules:\n"
            "- For chart_title: use chart_index from the provided charts list\n"
            "- Formulas start with = (e.g. =SUM(A1:A10))\n"
            "- You may combine multiple operations in one array\n"
            "- Return [] only if the instruction is truly impossible\n"
            "Return ONLY JSON array."
        ),
        messages=[{"role": "user", "content": f"Instruction: {instruction}\n\nFile structure:\n{json.dumps(file_structure, ensure_ascii=False)}"}]
    )

    print(f"[EXCEL EDIT DEBUG] Claude raw: {analysis_response.content[0].text}")

    raw = analysis_response.content[0].text.strip()
    if raw.startswith("```"):
        raw = raw.split("```")[1]
        if raw.startswith("json"):
            raw = raw[4:]

    try:
        ops = json.loads(raw.strip())
    except json.JSONDecodeError:
        logger.warning(f"targeted_excel_edit: invalid JSON from Claude: {raw[:200]}")
        return None, None, "Չհաջողվեց հասկանալ ինչ փոփոխություններ կատարել։ Խնդրեմ ավելի կոնկրետ նկարագրիր։"

    if not ops:
        return None, None, "Չհաջողվեց որոշել ինչ փոփոխություններ կատարել։ Խնդրեմ ավելի կոնկրետ նկարագրիր։"

    # ── Apply all operations ──
    wb = openpyxl.load_workbook(source_file_path)
    applied = []
    skipped = []

    for op in ops:
        op_type = op.get("type", "cell")
        sheet_name = op.get("sheet")

        if sheet_name and sheet_name not in wb.sheetnames:
            skipped.append(op)
            logger.warning(f"targeted_excel_edit: sheet not found: {sheet_name}")
            continue

        ws = wb[sheet_name] if sheet_name else wb.active

        try:
            if op_type == "cell":
                cell = op.get("cell")
                value = op.get("value")
                if cell and value is not None:
                    ws[cell] = value
                    applied.append(op)

            elif op_type == "fill":
                cell = op.get("cell")
                color = op.get("color", "").lstrip("#")
                if cell and color:
                    ws[cell].fill = PatternFill(start_color=color, end_color=color, fill_type="solid")
                    applied.append(op)

            elif op_type == "font":
                cell = op.get("cell")
                if cell:
                    existing = ws[cell].font
                    bold = op.get("bold", existing.bold)
                    italic = op.get("italic", existing.italic)
                    size = op.get("size", existing.size)
                    fcolor = op.get("color", "").lstrip("#") or None
                    ws[cell].font = Font(
                        bold=bold,
                        italic=italic,
                        size=size,
                        color=fcolor if fcolor else existing.color,
                    )
                    applied.append(op)

            elif op_type == "chart_title":
                chart_index = op.get("chart_index", 0)
                title = op.get("title", "")
                charts = getattr(ws, '_charts', [])
                if chart_index < len(charts):
                    charts[chart_index].title = title
                    applied.append(op)
                else:
                    skipped.append(op)
                    logger.warning(f"targeted_excel_edit: chart_index {chart_index} out of range")

            elif op_type == "chart_fill":
                chart_index = op.get("chart_index", 0)
                color = op.get("color", "").lstrip("#")
                charts = getattr(ws, '_charts', [])
                if chart_index < len(charts) and color:
                    from openpyxl.drawing.fill import PatternFillProperties
                    from openpyxl.chart.data_source import NumDataSource
                    chart = charts[chart_index]
                    try:
                        from openpyxl.drawing.spreadsheet_drawing import SpreadsheetDrawing
                        from openpyxl.chart._chart import AxDataSource
                        if hasattr(chart, 'plot_area') and hasattr(chart.plot_area, 'spPr'):
                            from openpyxl.drawing.fill import SolidColorFillProperties
                            chart.plot_area.spPr.solidFill = color
                    except Exception:
                        pass
                    applied.append(op)
                else:
                    skipped.append(op)

            elif op_type == "merge":
                cell_range = op.get("range")
                if cell_range:
                    ws.merge_cells(cell_range)
                    applied.append(op)

            elif op_type == "unmerge":
                cell_range = op.get("range")
                if cell_range:
                    ws.unmerge_cells(cell_range)
                    applied.append(op)

            elif op_type == "row_height":
                row = op.get("row")
                height = op.get("height")
                if row and height:
                    ws.row_dimensions[int(row)].height = float(height)
                    applied.append(op)

            elif op_type == "col_width":
                col = op.get("col", "").lstrip("#")
                width = op.get("width")
                if col and width:
                    ws.column_dimensions[col.upper()].width = float(width)
                    applied.append(op)

            else:
                skipped.append(op)
                logger.warning(f"targeted_excel_edit: unknown op type: {op_type}")

        except Exception as apply_err:
            skipped.append(op)
            logger.warning(f"targeted_excel_edit: failed to apply {op}: {apply_err}")

    if not applied:
        return None, None, "Փոփոխություններ կատարել չհաջողվեց։ Հնարավոր է ֆայլի կառուցվածքը չի համապատասխանում հրահանգին։"

    # ── Save ──
    file_id = str(uuid.uuid4())
    output_path = str(UPLOAD_DIR / f"excel_{file_id}.xlsx")
    wb.save(output_path)
    print(f"[EXCEL EDIT DEBUG] file_id={file_id}, applied={len(applied)}, skipped={len(skipped)}")

    preview = {
        "columns": ["type", "sheet", "detail"],
        "rows": [
            [e.get("type"), e.get("sheet"), e.get("cell") or e.get("range") or f"chart[{e.get('chart_index',0)}]"]
            for e in applied
        ],
        "total_rows": len(applied),
        "message": f"Applied {len(applied)} operation(s)." + (f" Skipped {len(skipped)}." if skipped else ""),
    }
    summary = f"Կատարվեց {len(applied)} փոփոխություն։" + (f" Չկատարվեց {len(skipped)}։" if skipped else "")
    return file_id, preview, summary


async def scratch_generate_excel(
    message_content: str,
    ai_response_text: str,
    claude_client,
) -> Tuple[Optional[str], Optional[dict], str]:
    """Generate an Excel file from scratch using Claude.

    Uses the user's message and the AI's text response as context to produce
    structured data, then writes it to an .xlsx file with openpyxl.

    Returns (file_id, preview, message_text) or (None, None, error_text).
    """
    import openpyxl
    from openpyxl.styles import PatternFill, Font, Alignment

    print(f"[EXCEL SCRATCH] generating from scratch for: {message_content[:80]}")

    # Ask Claude to produce structured sheet data
    gen_response = await claude_client.messages.create(
        model="claude-sonnet-4-6",
        max_tokens=4096,
        system=(
            "You are an Excel file generator. The user wants to create an Excel file from scratch.\n"
            "The user's instruction and any relevant context from the conversation are provided below.\n"
            "Respond ONLY with a valid JSON object — no markdown fences, no explanation.\n\n"
            "Required format:\n"
            '{\n'
            '  "filename": "report.xlsx",\n'
            '  "sheets": [\n'
            '    {\n'
            '      "name": "Sheet1",\n'
            '      "headers": ["Column A", "Column B", "Column C"],\n'
            '      "rows": [\n'
            '        ["value1", "value2", 100],\n'
            '        ["value3", "value4", 200]\n'
            '      ]\n'
            '    }\n'
            '  ],\n'
            '  "message": "Brief description of what was generated (in same language as user)"\n'
            '}\n\n'
            "Rules:\n"
            "- Use the conversation context to populate realistic data\n"
            "- Headers must match the data context (use user's language for column names)\n"
            "- Include at least 3-5 data rows (more if the context provides them)\n"
            "- Numbers should be actual numbers (not strings)\n"
            "- Dates as strings in YYYY-MM-DD format\n"
            "- Multiple sheets are allowed when it makes sense\n"
            "- Return ONLY the JSON object, no extra text"
        ),
        messages=[{
            "role": "user",
            "content": (
                f"User request: {message_content}\n\n"
                f"Conversation context / AI response:\n{ai_response_text[:3000]}"
            )
        }]
    )

    raw = gen_response.content[0].text.strip()
    print(f"[EXCEL SCRATCH] Claude raw (first 300): {raw[:300]}")

    # Strip markdown fences if present
    if raw.startswith("```"):
        raw = raw.split("```")[1]
        if raw.startswith("json"):
            raw = raw[4:]
        raw = raw.strip()

    try:
        sheet_data = json.loads(raw)
    except json.JSONDecodeError as e:
        logger.error(f"scratch_generate_excel: JSON parse error: {e} | raw: {raw[:200]}")
        return None, None, "Excel ֆայլ ստեղծել չհաջողվեց։ Claude-ը անճիշտ ձևաչափ վերադարձրեց։"

    # Claude might return a list instead of an object
    if not isinstance(sheet_data, dict):
        logger.error(f"scratch_generate_excel: expected dict, got {type(sheet_data)}")
        return None, None, "Excel ֆայլ ստեղծել չհաջողվեց — անսպասելի ձևաչափ։"

    sheets = sheet_data.get("sheets", [])
    if not sheets:
        return None, None, "Excel ֆայլ ստեղծել չհաջողվեց — data չստացվեց։"

    raw_filename = sheet_data.get("filename") or "report"
    base = raw_filename.rsplit(".", 1)[0] if "." in raw_filename else raw_filename
    filename = (base.strip() or "report") + ".xlsx"

    # ── Build the workbook ──
    wb = openpyxl.Workbook()
    # Remove the default empty sheet openpyxl always creates
    if wb.active is not None:
        wb.remove(wb.active)

    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=11)
    header_align = Alignment(horizontal="center", vertical="center")

    all_preview_rows = []
    all_preview_cols = []
    total_rows = 0
    used_sheet_names: set = set()

    for sheet_def in sheets:
        raw_name = str(sheet_def.get("name") or "Sheet")[:31]

        # Deduplicate sheet names — openpyxl raises if two sheets share a name
        sheet_name = raw_name
        counter = 2
        while sheet_name in used_sheet_names:
            suffix = f" ({counter})"
            sheet_name = raw_name[:31 - len(suffix)] + suffix
            counter += 1
        used_sheet_names.add(sheet_name)

        headers = sheet_def.get("headers") or []
        rows = sheet_def.get("rows") or []

        ws = wb.create_sheet(title=sheet_name)

        # Write headers with formatting
        for col_idx, header in enumerate(headers, start=1):
            cell = ws.cell(row=1, column=col_idx, value=str(header))
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = header_align
            ws.column_dimensions[cell.column_letter].width = max(15, len(str(header)) + 4)

        # Write data rows
        for row_idx, row_data in enumerate(rows, start=2):
            if not isinstance(row_data, (list, tuple)):
                continue
            for col_idx, value in enumerate(row_data, start=1):
                ws.cell(row=row_idx, column=col_idx, value=_sanitize_value(value))

        if headers:
            ws.row_dimensions[1].height = 20

        total_rows += len(rows)

        if not all_preview_cols and headers:
            all_preview_cols = [str(h) for h in headers]
        if len(all_preview_rows) < 5 and rows:
            all_preview_rows.extend(
                [[_sanitize_value(v) for v in r] for r in rows[:5 - len(all_preview_rows)]]
            )

    # ── Save ──
    file_id = str(uuid.uuid4())
    output_path = str(UPLOAD_DIR / f"excel_{file_id}.xlsx")
    wb.save(output_path)
    print(f"[EXCEL SCRATCH] saved: {output_path}, sheets={len(sheets)}, total_rows={total_rows}")

    preview = {
        "columns": all_preview_cols,
        "rows": all_preview_rows,
        "total_rows": total_rows,
        "message": sheet_data.get("message", f"Generated {filename}"),
    }
    message_text = sheet_data.get("message", f"Excel ֆայլը պատրաստ է — {total_rows} տող, {len(sheets)} sheet։")
    return file_id, preview, message_text


async def maybe_generate_excel(
    db,
    chat_id: str,
    project_id: str,
    active_source_ids: list,
    message_content: str,
    claude_client,
    current_response_text: str,
    temp_file_path: Optional[str] = None,
) -> Tuple[Optional[str], Optional[dict], str, bool]:
    """
    Attempt to generate an Excel file if conditions are met.

    Decision tree:
    1. If no Excel trigger in message → return text as-is.
    2. If a temp/source Excel file exists → targeted_edit (modify existing file).
    3. If no file exists → scratch_generate_excel (create from scratch using AI context).

    Returns (excel_file_id, excel_preview, response_text, is_clarification).
    """
    if not is_excel_trigger(message_content):
        print(f"[EXCEL] no trigger found in message: {message_content[:80]}")
        return None, None, current_response_text, False

    effective_source_ids = active_source_ids or []
    print(f"[EXCEL] start: project_id={project_id}, active_source_ids_count={len(effective_source_ids)}, has_temp={bool(temp_file_path)}")

    try:
        # ── Step 1: Resolve existing file (temp upload or project source) ──
        actual_file_path = None
        actual_ext = None
        source_name = "result.xlsx"

        if temp_file_path and Path(temp_file_path).exists():
            actual_file_path = Path(temp_file_path)
            actual_ext = actual_file_path.suffix.lstrip(".").lower()
            source_name = actual_file_path.name.split("_", 1)[-1]
            print(f"[EXCEL] using temp file: {source_name}")

        elif project_id or effective_source_ids:
            # Search project/active sources for an Excel file
            id_filter = (
                {"id": {"$in": effective_source_ids}}
                if effective_source_ids
                else {"projectId": project_id}
            )

            excel_source = await db.sources.find_one(
                {
                    **id_filter,
                    "mimeType": {"$in": [
                        "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                        "application/vnd.ms-excel",
                    ]}
                },
                {"_id": 0}
            )
            if not excel_source:
                excel_source = await db.sources.find_one(
                    {**id_filter, "mimeType": {"$in": ["text/csv", "application/csv"]}},
                    {"_id": 0}
                )
            if not excel_source:
                all_sources = await db.sources.find(id_filter, {"_id": 0}).to_list(100)
                for s in all_sources:
                    sp = (s.get("storagePath") or "").lower()
                    on = (s.get("originalName") or "").lower()
                    if sp.endswith((".xlsx", ".xls", ".csv")) or on.endswith((".xlsx", ".xls", ".csv")):
                        excel_source = s
                        print(f"[EXCEL] found source by extension fallback: {on}")
                        break

            if excel_source and excel_source.get("storagePath"):
                candidate = UPLOAD_DIR / excel_source["storagePath"]
                if candidate.exists():
                    actual_file_path = candidate
                    actual_ext = excel_source["storagePath"].rsplit(".", 1)[-1].lower()
                    source_name = excel_source.get("originalName", "file")
                    print(f"[EXCEL] resolved source: {source_name}, exists=True")
                else:
                    print(f"[EXCEL] source found in DB but missing on disk: {candidate}")

        # ── Step 2a: Existing xlsx/xls file → targeted edit ──
        # CSVs are skipped here because targeted_edit uses openpyxl which can't
        # open a CSV. They fall through to scratch generation instead, which reads
        # the CSV content via the AI response context and produces a proper .xlsx.
        if actual_file_path and actual_file_path.exists() and actual_ext in ("xlsx", "xls", "xlsm"):
            print(f"[EXCEL] routing to targeted_edit (ext={actual_ext})")
            file_id, preview, text = await targeted_excel_edit(
                str(actual_file_path), message_content, claude_client
            )
            if file_id:
                await _persist_excel_to_db(
                    db, file_id, str(UPLOAD_DIR / f"excel_{file_id}.xlsx"), source_name
                )
                return file_id, preview, text, False

            # targeted_edit returned no ops — surface the error as clarification
            clarif_text = (
                text if text else
                "Չհասկացա ինչ փոփոխություն կատարել։ Խնդրեմ կոնկրետ նկարագրիր — "
                "օրինակ՝ «A2 բջիջում գրիր 100» կամ «Price սյունակի բոլոր արժեքները բազմապատկիր 1.2-ով»։"
            )
            print(f"[EXCEL] targeted_edit returned no ops")
            return None, None, clarif_text, True

        # ── Step 2b: No file → generate from scratch ──
        print(f"[EXCEL] no existing file found — routing to scratch generation")
        file_id, preview, text = await scratch_generate_excel(
            message_content=message_content,
            ai_response_text=current_response_text,
            claude_client=claude_client,
        )
        if file_id:
            await _persist_excel_to_db(db, file_id, str(UPLOAD_DIR / f"excel_{file_id}.xlsx"), "generated.xlsx")
            return file_id, preview, text, False

        # scratch generation failed — return whatever error text it gave
        return None, None, text or current_response_text, False

    except Exception as excel_err:
        logger.error(f"Excel generation error: {excel_err}", exc_info=True)
        return None, None, current_response_text, False
